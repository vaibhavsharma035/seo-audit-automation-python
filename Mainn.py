import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import urllib3

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Excel file
input_file = "Copy of Rainwater, Holt & Sexton - SAT & Ops Center (1).xlsx"

# Sheet name (IMPORTANT)
sheet_name = "Ops Center"

# Required columns
required_columns = [
    "42Works - Staging Site Link",
    "New H1",
    "Live Title Tag",
    "Live Meta Description"
]

# Read ONLY Ops Center sheet
df = pd.read_excel(input_file, sheet_name=sheet_name, header=1)

# Clean column names
df.columns = df.columns.astype(str).str.strip()

# Keep ONLY required columns
df = df[required_columns]

# Rename for easy use
df.columns = ["URL", "Expected_H1", "Expected_Title", "Expected_Meta"]

# Remove empty URLs
df = df[df["URL"].notna()].reset_index(drop=True)

# Result columns
df["Status_Code"] = ""
df["Actual_H1"] = ""
df["H1_Count"] = 0
df["H1_Result"] = ""

df["Actual_Title"] = ""
df["Title_Result"] = ""

df["Actual_Meta"] = ""
df["Meta_Result"] = ""

# Request session
session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
})

def clean_text(text):
    if text is None or pd.isna(text):
        return ""
    return " ".join(str(text).replace("\n", " ").replace("\r", " ").split()).strip()

for index, row in df.iterrows():

    url = row["URL"]

    expected_h1 = clean_text(row["Expected_H1"])
    expected_title = clean_text(row["Expected_Title"])
    expected_meta = clean_text(row["Expected_Meta"])

    print("Checking:", url)

    try:
        response = session.get(url, timeout=15, verify=False)  # FIXED
        status_code = response.status_code

        soup = BeautifulSoup(response.text, "html.parser")

        # H1
        h1_tags = soup.find_all("h1")
        h1_count = len(h1_tags)

        actual_h1 = clean_text(h1_tags[0].get_text()) if h1_count > 0 else ""

        # Title
        actual_title = clean_text(soup.title.get_text()) if soup.title else ""

        # Meta description
        meta_tag = soup.find("meta", attrs={"name": "description"})

        if not meta_tag:
            meta_tag = soup.find("meta", attrs={"property": "og:description"})

        actual_meta = clean_text(meta_tag.get("content")) if meta_tag and meta_tag.get("content") else ""

    except Exception:

        print("Error loading:", url)

        status_code = "ERROR"
        actual_h1 = ""
        actual_title = ""
        actual_meta = ""
        h1_count = 0

    # Save extracted values
    df.at[index, "Status_Code"] = str(status_code)
    df.at[index, "Actual_H1"] = actual_h1
    df.at[index, "H1_Count"] = h1_count
    df.at[index, "Actual_Title"] = actual_title
    df.at[index, "Actual_Meta"] = actual_meta

    # ---------------- H1 Result ----------------

    if expected_h1 == "":
        h1_result = "NO EXPECTED H1"
    elif actual_h1 == "":
        h1_result = "H1 MISSING"
    elif h1_count > 1:
        h1_result = "MULTIPLE H1"
    elif expected_h1.lower() == actual_h1.lower():
        h1_result = "PASS"
    else:
        h1_result = "FAIL"

    # ---------------- Title Result ----------------

    if expected_title == "":
        title_result = "NO EXPECTED TITLE"
    elif actual_title == "":
        title_result = "TITLE MISSING"
    elif expected_title.lower() == actual_title.lower():
        title_result = "PASS"
    else:
        title_result = "FAIL"

    # ---------------- Meta Result ----------------

    if expected_meta == "":
        meta_result = "NO EXPECTED META"
    elif actual_meta == "":
        meta_result = "META MISSING"
    elif expected_meta.lower() == actual_meta.lower():
        meta_result = "PASS"
    else:
        meta_result = "FAIL"

    df.at[index, "H1_Result"] = h1_result
    df.at[index, "Title_Result"] = title_result
    df.at[index, "Meta_Result"] = meta_result

    # Delay to avoid blocking
    time.sleep(1)


# Save output
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = f"seo_results_{timestamp}.xlsx"

df.to_excel(output_file, index=False)


# -------- Highlight FAIL cells --------

wb = load_workbook(output_file)
ws = wb.active

fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

columns_to_check = ["H1_Result", "Title_Result", "Meta_Result"]
col_map = {}

for cell in ws[1]:
    if cell.value in columns_to_check:
        col_map[cell.value] = cell.column

for row in range(2, ws.max_row + 1):
    for col in col_map.values():

        cell = ws.cell(row=row, column=col)

        if str(cell.value).strip().upper() == "FAIL":
            cell.fill = fail_fill

wb.save(output_file)

print("\nSEO report generated:", output_file)